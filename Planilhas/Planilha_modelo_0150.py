import pandas as pd
from openpyxl import load_workbook
import re

# Mapeamento de colunas
column_mapping = {
    "Código Participante": "Código Participante",
    "Nome Participante": "Nome Participante",
    "País": "País",
    "CNPJ": "CNPJ",
    "CPF Participante": "",
    "Inscrição Estadual Participante": "",
    "Código Município": "Código Município",
    "SUFRAMA": "",
    "Endereço": "Endereço",
    "Número": "Número",
    "Complemento": "",
    "Bairro": "Bairro"
}


# Caminhos para os arquivos
origem_path = "C:\\Users\\Axxen\\Desktop\\RETIFICAÇÃ\\EFD - ICMS-IPI - 0150 - Cadastro do Participante difripar.xlsx"
destino_path = "C:\\Users\\Axxen\\Desktop\\RETIFICAÇÃ\\Planilha_modelo_0150_atualizada.xlsx"

# Carregar a planilha de origem
df_origem = pd.read_excel(origem_path, sheet_name='Planilha', dtype=str)

# Criar um DataFrame vazio para a planilha de destino
df_destino = pd.DataFrame()

# Preencher o DataFrame de destino com base no mapeamento
for destino_col, origem_col in column_mapping.items():
    if origem_col:
        df_destino[destino_col] = df_origem.get(origem_col, "")
    else:
        df_destino[destino_col] = ""

# Substituir valores vazios por "0"
df_destino.fillna("0", inplace=True)

# Carregar a planilha de destino existente
book = load_workbook(destino_path)
sheet = book['Planilha1']  # Certifique-se de que a planilha correta está selecionada

# Localizar a linha da tabela que contém os cabeçalhos
linha_cabecalho = 2

# Encontrar a última linha preenchida na tabela a partir do cabeçalho
start_row = linha_cabecalho + 1
while sheet.cell(row=start_row, column=1).value is not None:
    start_row += 1

# Adicionar o DataFrame à planilha 'Planilha1' sem incluir o cabeçalho
for r_idx, row in df_destino.iterrows():
    for c_idx, value in enumerate(row):
        sheet.cell(row=start_row + r_idx, column=c_idx + 1, value=value)

# Salvar o arquivo
book.save(destino_path)

print("Dados transferidos com sucesso!")
