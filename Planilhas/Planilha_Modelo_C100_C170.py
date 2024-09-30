import pandas as pd
from openpyxl import load_workbook
import re

# Mapeamento de colunas
column_mapping = {
    "**IND_OPER_C100": "Tipo Operação",
    "**IND_EMIT_C100": "Indicador Emitente",
    "COD_PART_C100": "Código Participante",
    "COD_MOD": "Modelo",
    "**COD_SIT_C100": "Situação",
    "SER_C100": "Série",
    "**NUM_DOC_C100": "Número Documento",
    "CHV_NFSE_C100": "Chave NF-e",
    "DT_DOC_C100": "Data Documento",
    "DT_E_S_C100": "Data Entrada/Saída",
    "**VL_DOC_C100": "Vlr Documento",
    "**IND_PGTO_C100": "Indicador Pagamento",
    "VL_DESC_C100": "Vlr Desconto NF",
    "VL_ABAT_NT": "Vlr Abatimento NT",
    "VL_MERC": "Vlr Mercadoria",
    "IND_FRT": "Indicador Frete",
    "VL_FRT": "Vlr Frete",
    "VL_SEG": "Vlr Seguro",
    "VL_OUT_DA": "Vlr Outras DA",
    "VL_BC_ICMS": "Vlr Base Cálculo ICMS - C100",
    "VL_ICMS": "Vlr ICMS - C100",
    "VL_BC_ICMS_ST": "Vlr Base Cálculo ICMS ST - C100",
    "VL_ICMS_ST_C100": "Vlr ICMS ST - C100",
    "VL_IPI_C100": "Vlr IPI - C100",
    "VL_PIS": "Vlr PIS - C100",
    "VL_COFINS": "Vlr Cofins - C100",
    "VL_PIS_ST": "Vlr PIS ST - C100",
    "VL_COFINS_ST": "Vlr Cofins - C100",
    "NUM_ITEM": "Número Item",
    "COD_ITEM": "Código Item",
    "DESCR_COMPL": "Descrição Item",
    "QTD": "Qtde Item",
    "UNID": "Unidade Medida",
    "VL_ITEM": "Vlr Item",
    "VL_DESC": "Vlr Desconto Item",
    "IND_MOV": "Indicador Movimento Item",
    "CST_ICMS": "CST ICMS - C170/C190",
    "CFOP": "CFOP",
    "COD_NAT": "",
    "VL_BC_ICMS_C170": "Vlr Base Cálculo ICMS - C170/C190",
    "ALIQ_ICMS": "Alíquota ICMS - C170/C190",
    "VL_ICMS_C170": "Vlr ICMS - C170/C190",
    "VL_BC_ICMS_ST_C170": "Vlr Base Cálculo ICMS ST - C170/C190",
    "ALIQ_ST": "Alíquota ICMS ST - C170",
    "VL_ICMS_ST": "Vlr ICMS ST - C170/C190",
    "IND_APUR": "Indicador Apuração IPI",
    "CST_IPI": "CST IPI",
    "COD_ENQ": "",
    "VL_BC_IPI": "Vlr Base Cálculo IPI",
    "ALIQ_IPI": "Alíquota IPI",
    "VL_IPI": "Vlr IPI - C170/C190",
    "CST_PIS": "CST PIS/Cofins",
    "VL_BC_PIS": "Vlr Base Cálculo PIS",
    "ALIQ_PIS_PERC": "Alíquota PIS",
    "QUANT_BC_PIS": "",
    "ALIQ_PIS_REAIS": "",
    "VL_PIS_C170": "Vlr PIS",
    "CST_COFINS": "CST PIS/Cofins",
    "VL_BC_COFINS": "Vlr Base Cálculo Cofins",
    "ALIQ_COFINS_PERC": "Alíquota Cofins",
    "QUANT_BC_COFINS": "",
    "ALIQ_COFINS_REAIS": "",
    "VL_COFINS_C170": "Vlr Cofins",
    "COD_CTA": "Conta Contábil",
    "INFO_COMPL": ""
}

# Caminhos para os arquivos
origem_path = "C:\\Users\\Axxen\\Desktop\\SCRIPT LEO\\tarefa\M SAM.xlsx"
destino_path = "C:\\Users\\Axxen\\Desktop\\SCRIPT LEO\\tarefa\\PLANILHA_MODELO_REVENDA.xlsx"

# Carregar a planilha de origem
df_origem = pd.read_excel(origem_path, sheet_name='MANUT VEÍCULO', dtype=str)

# Criar um DataFrame vazio para a planilha de destino
df_destino = pd.DataFrame()

# Preencher o DataFrame de destino com base no mapeamento
for destino_col, origem_col in column_mapping.items():
    if origem_col:
        df_destino[destino_col] = df_origem.get(origem_col, "")
    else:
        df_destino[destino_col] = ""

# Aplicar as transformações solicitadas
# Remover letras e manter apenas números
df_destino["**IND_OPER_C100"] = df_destino["**IND_OPER_C100"].str.extract('(\d+)')
df_destino["**IND_EMIT_C100"] = df_destino["**IND_EMIT_C100"].str.extract('(\d+)')
df_destino["**IND_PGTO_C100"] = df_destino["**IND_PGTO_C100"].str.extract('(\d+)')
df_destino["IND_APUR"] = df_destino["IND_APUR"].str.extract('(\d+)')
df_destino["IND_FRT"] = df_destino["IND_FRT"].str.extract('(\d+)')
df_destino["IND_MOV"] = df_destino["IND_MOV"].str.extract('(\d+)')

# Remover barras das datas
df_destino["DT_DOC_C100"] = df_destino["DT_DOC_C100"].str.replace("/", "")
df_destino["DT_E_S_C100"] = df_destino["DT_E_S_C100"].str.replace("/", "")

df_destino = df_destino.applymap(lambda x: x.replace('.', ',') if isinstance(x, str) else x)

# Substituir valores vazios por "0" nas colunas especificadas
#df_destino["COD_ENQ"] = df_destino["COD_ENQ"].replace("0", "")
df_destino["COD_NAT"] = df_destino["COD_NAT"].replace("", "0")
#df_destino["ALIQ_PIS_REAIS"] = df_destino["ALIQ_PIS_REAIS"].replace("", "0")
#df_destino["ALIQ_COFINS_REAIS"] = df_destino["ALIQ_COFINS_REAIS"].replace("", "0")
#df_destino["INFO_COMPL"] = df_destino["INFO_COMPL"].replace("", "0")

# Preencher valores vazios com "0"
df_destino.fillna("0", inplace=True)

# Carregar a planilha de destino existente
book = load_workbook(destino_path)
sheet = book['MANUT VEÍCULO']

# Localizar a linha da tabela que contém os cabeçalhos
linha_cabecalho = 2

# Encontrar a última linha preenchida na tabela a partir do cabeçalho
start_row = linha_cabecalho + 1
while sheet.cell(row=start_row, column=1).value is not None:
    start_row += 1

# Adicionar o DataFrame à planilha 'Planilha2' sem incluir o cabeçalho
for r_idx, row in df_destino.iterrows():
    for c_idx, value in enumerate(row):
        sheet.cell(row=start_row + r_idx, column=c_idx + 1, value=value)

# Salvar o arquivo
book.save(destino_path)

print("Dados transferidos com sucesso!")
