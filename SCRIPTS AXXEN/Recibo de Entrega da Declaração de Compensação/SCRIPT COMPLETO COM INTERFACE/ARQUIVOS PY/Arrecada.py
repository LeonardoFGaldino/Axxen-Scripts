import os
import fitz
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog, Entry
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

pasta_pdf = ""
pasta_destino = ""


def extrair_valores_cp(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        texto = ""
        for pagina in doc:
            texto += pagina.get_text()

        padrao = re.compile(
            r'CP TERCEIROS\s+([\d\.]+,\d+)'
            r'|SALÁRIO EDUCAÇÃO\n([\d.,]+)'
            r'|INCRA\n([\d.,]+)'
            r'|SENAI\n([\d.,]+)'
            r'|SESI\n([\d.,]+)'
            r'|SEBRAE/APEX/ABDI\n([\d.,]+)',
            re.IGNORECASE
        )

        valores_encontrados = padrao.findall(texto)

        cp_terceiros = ""
        salario_educacao = ""
        incra = ""
        senai = ""
        sesi = ""
        abdi = ""



        for match in valores_encontrados:
            if match[0]:
                cp_terceiros = float(match[0].replace(".", "").replace(",", "."))
            elif match[1]:
                salario_educacao = float(match[1].replace(".", "").replace(",", "."))
            elif match[2]:
                incra = float(match[2].replace(".", "").replace(",", "."))
            elif match[3]:
                senai = float(match[3].replace(".", "").replace(",", "."))
            elif match[4]:
                sesi = float(match[4].replace(".", "").replace(",", "."))
            elif match[5]:
                abdi = float(match[5].replace(".", "").replace(",", "."))



        return cp_terceiros, salario_educacao, incra, senai, senai, sesi, abdi
    except Exception as e:
        print(f"Erro ao extrair informações do PDF: {str(e)}")
        return 0.0, 0.0, 0.0


def selecionar_pasta():
    global pasta_pdf
    pasta_pdf = filedialog.askdirectory()
    pasta_label.config(text=f"Pasta com PDFs: {pasta_pdf}")


def selecionar_pasta_destino():
    global pasta_destino
    pasta_destino = filedialog.askdirectory()
    pasta_destino_label.config(text=f"Pasta de Destino: {pasta_destino}")


def extrair_e_salvar():
    informacoes_pdf = []

    for arquivo in os.listdir(pasta_pdf):
        if arquivo.endswith('.pdf'):
            pdf_path = os.path.join(pasta_pdf, arquivo)
            cp_terceiros, salario_educacao, incra, senai, senai, sesi, abdi = extrair_valores_cp(
                pdf_path)
            informacoes_pdf.append((arquivo, cp_terceiros, salario_educacao, incra, senai, senai, sesi, abdi))


    df = pd.DataFrame(
        informacoes_pdf,
        columns=["Nome do Arquivo", "CP TERCEIROS", "1170 CP TERCEIROS - SALÁRIO EDUCAÇÃO", "1176 CP TERCEIROS - INCRA",
                 "1181 CP TERCEIROS - SENAI", "1184 CP TERCEIROS - SESI", "1200  CP TERCEIROS - SEBRAE",
                 "1234 Coluna Extra"]
    )

    colunas_numericas = df.columns[1:]

    df[colunas_numericas] = df[colunas_numericas].apply(pd.to_numeric, errors='coerce').fillna(0)


    df['Soma Total'] = df[colunas_numericas].sum(axis=1)

    total_geral = sum(list(df['Soma Total']))

    qnty_cols = len(df.columns)
    ultima_linha = [''] * qnty_cols
    ultima_linha[-1] = total_geral
    df.loc['Total Geral'] = ultima_linha
    df[colunas_numericas] = df[colunas_numericas].apply(lambda valor: valor.replace(0, ''))



    if not pasta_destino:
        resultado_label.config(text="Por favor, selecione uma pasta de destino.")
        return

    nome_arquivo = nome_arquivo_entry.get().strip()

    if not nome_arquivo:
        resultado_label.config(text="Por favor, insira um nome de arquivo.")
        return

    if not nome_arquivo.endswith(".xlsx"):
        nome_arquivo += ".xlsx"

    excel_filename = os.path.join(pasta_destino, nome_arquivo)

    pasta_completa = os.path.abspath(pasta_pdf)
    df['Nome do Arquivo'] = df.apply(lambda row: f'{os.path.join(pasta_completa, row["Nome do Arquivo"])}', axis=1)

    workbook = Workbook()
    ws = workbook.active

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.hyperlink = cell.value

    workbook.save(excel_filename)
    resultado_label.config(text=f'Valores extraídos e salvos em "{excel_filename}"')



root = tk.Tk()
root.title("Extração de Valores de PDF")

selecionar_button = tk.Button(root, text="Selecionar Pasta com PDFs", command=selecionar_pasta)
selecionar_destino_button = tk.Button(root, text="Selecionar Pasta de Destino", command=selecionar_pasta_destino)
extrair_button = tk.Button(root, text="Extrair e Salvar", command=extrair_e_salvar, width=50)
pasta_label = tk.Label(root, text="Pasta com PDFs não selecionada.")
pasta_destino_label = tk.Label(root, text="Pasta de Destino não selecionada.")
resultado_label = tk.Label(root, text="")

nome_arquivo_label = tk.Label(root, text="Nome do arquivo de saída:")
nome_arquivo_entry = Entry(root)

selecionar_button.pack()
selecionar_destino_button.pack()
pasta_label.pack()
pasta_destino_label.pack()
nome_arquivo_label.pack()
nome_arquivo_entry.pack()
extrair_button.pack()
resultado_label.pack()

root.mainloop()
