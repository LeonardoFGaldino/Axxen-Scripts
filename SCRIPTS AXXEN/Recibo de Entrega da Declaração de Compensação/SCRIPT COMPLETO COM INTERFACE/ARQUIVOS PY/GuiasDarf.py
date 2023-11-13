import fitz
import os
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from glob import glob

files = glob('*pdf')
files = [f for f in files if not f.startswith('~$')]


# Função para extrair valores dos PDFs
def extrair_valores_cnpj_documento_e_razao_social(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        texto = ""
        for pagina in doc:
            texto += pagina.get_text()

        cnpj_pattern = r'\d{2}\.\d{3}\.\d{3}\/\d{4}\-\d{2}'
        match_cnpj = re.search(cnpj_pattern, texto) or ''
        if match_cnpj:
            match_cnpj = match_cnpj.group()

        # Encontrar a posição do CNPJ no texto
        posicao_cnpj = texto.find(match_cnpj)

        # Extrair a Razão Social após o CNPJ
        razao_social_pattern = r'{}\s+(.+?)\n'.format(re.escape(match_cnpj))
        match_razao_social = re.search(razao_social_pattern, texto) or ''
        if match_razao_social:
            match_razao_social = match_razao_social.group(1)
        else:
            match_razao_social = "Razão Social não encontrada"

        # Regex para um número fictício de documento com 10 dígitos (altere conforme necessário)
        documento_pattern = r'\s+([\d.]+,\d+)'
        match_documento = re.search(documento_pattern, texto) or ''
        if match_documento:
            match_documento = match_documento.group()

        # Extrair o código usando uma expressão regular
        codigo_pattern = r'(?i)Arrecadação\s+(\d+)'
        match_codigo = re.search(codigo_pattern, texto)

        print("Texto extraído do PDF:")
        print(texto)

        if match_codigo:
            match_codigo = match_codigo.group(1)
        else:
            match_codigo = "Código não encontrado"

        print("Código encontrado:")
        print(match_codigo)

        return [match_cnpj, float(match_documento.replace('.', '').replace(',', '.')), match_razao_social,
                int(match_codigo)]
    except Exception as e:
        print(f"Erro ao extrair informações do PDF: {str(e)}")
        return ["", 0.0, "Razão Social não encontrada", 0]


# Função para iniciar a extração quando o botão for clicado
def iniciar_extracao():
    global pasta_pdf
    pasta_pdf = pasta_entry.get()
    pasta_label.config(text=f"Pasta com PDFs: {pasta_pdf}")

    informacoes_pdf.clear()
    for arquivo in os.listdir(pasta_pdf):
        if arquivo.endswith('.pdf') and not arquivo.startswith('~$'):
            pdf_path = os.path.join(pasta_pdf, arquivo)
            valores = extrair_valores_cnpj_documento_e_razao_social(pdf_path)
            informacoes_pdf.append([arquivo] + valores)

    # Criar um DataFrame
    colunas = ["Nome do Arquivo", "CNPJ", "Valor Total do Documento", "Razão Social", "Código"]
    df = pd.DataFrame(informacoes_pdf, columns=colunas)

    # Solicitar ao usuário o nome do arquivo Excel
    nome_arquivo_excel = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                      filetypes=[("Arquivos Excel", "*.xlsx")])

    if nome_arquivo_excel:
        # Salvar o DataFrame no arquivo Excel
        df.to_excel(nome_arquivo_excel, index=False, engine='openpyxl')

        # Carregar o arquivo Excel com openpyxl
        wb = load_workbook(nome_arquivo_excel)
        ws = wb.active

        # Adicionar hiperlinks aos nomes dos arquivos (alternativa)
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            cell = row[0]
            nome_arquivo = cell.value
            if nome_arquivo:
                caminho_completo = os.path.join(pasta_pdf, nome_arquivo)
                cell.hyperlink = caminho_completo
                cell.style = "Hyperlink"

        # Salvar o arquivo Excel atualizado
        wb.save(nome_arquivo_excel)

        resultado_label.config(text=f'Valores extraídos e salvos em "{nome_arquivo_excel}".')


# Iniciar a aplicação
root = tk.Tk()
root.title("Extração de CNPJ, Documento, Razão Social e Código de PDFs")

# Widgets
instrucoes_label = tk.Label(root, text="Informe o diretório:")
pasta_label = tk.Label(root, text="Pasta com PDFs não selecionada.")
pasta_entry = tk.Entry(root, width=50)
selecionar_button = tk.Button(root, text="Selecionar Pasta",
                              command=lambda: pasta_entry.insert(0, filedialog.askdirectory()))
iniciar_button = tk.Button(root, text="Iniciar Extração", command=iniciar_extracao)
resultado_label = tk.Label(root, text="")

# Layout dos widgets
instrucoes_label.pack()
pasta_entry.pack()
selecionar_button.pack()
pasta_label.pack()
iniciar_button.pack()
resultado_label.pack()

# Lista para armazenar informações dos PDFs
informacoes_pdf = []

# Executar a aplicação
root.mainloop()
