import fitz
import os
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl import load_workbook
from glob import glob

files = glob('*pdf')
files = [f for f in files if not f.startswith('~$')]

# Função para extrair valores dos PDFs
def extrair_valores_cnpj_razao_social_e_valor_total(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        texto = ""
        for pagina in doc:
            texto += pagina.get_text()

        print("Texto extraído:")
        print(texto)  # Adiciona esta linha para imprimir o texto extraído

        # Padrão para extrair CNPJ
        cnpj_pattern = r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}'
        match_cnpj = re.search(cnpj_pattern, texto)
        cnpj = match_cnpj.group() if match_cnpj else ''

        # Padrão para extrair Razão Social sem número de telefone (primeiro tipo de PDF)
        razao_social_pattern1 = r'NOME/TELEFONE\n(.+?)\n'
        match_razao_social1 = re.search(razao_social_pattern1, texto)
        razao_social1 = match_razao_social1.group(1).strip().split(' - ')[0] if match_razao_social1 else ''

        # Padrão para extrair Razão Social com CNPJ como prefixo (segundo tipo de PDF)
        razao_social_pattern2 = r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\n(.+?)\n'
        match_razao_social2 = re.search(razao_social_pattern2, texto)
        razao_social2 = match_razao_social2.group(1).strip() if match_razao_social2 else ''

        print("Padrão de regex para valor_total_pattern2:")
        print(valor_total_pattern2)  # Adiciona esta linha para imprimir o padrão de regex

        # Padrão para extrair Valor Total do primeiro tipo de PDF
        valor_total_pattern1 = r'VALOR TOTAL\s*([\d.,]+)'
        match_valor_total1 = re.search(valor_total_pattern1, texto)
        valor_total1 = match_valor_total1.group(1).replace(',', '') if match_valor_total1 else '0.0'

        # Padrão para extrair Valor Total do segundo tipo de PDF
        valor_total_pattern2 = r'Valor Total do Documento\s+([\d.,]+)'
        match_valor_total2 = re.search(valor_total_pattern2, texto, re.IGNORECASE)
        valor_total2 = match_valor_total2.group(1).replace('.', '').replace(',', '.') if match_valor_total2 else '0.0'

        print("Resultado do match_valor_total2:")
        print(match_valor_total2)  # Adiciona esta linha para imprimir o resultado do regex

        # Escolhe o Valor Total baseado nos padrões encontrados
        valor_total = valor_total1 if valor_total1 else valor_total2

        return cnpj, razao_social, float(valor_total.replace(',', '.'))

    except Exception as e:
        print(f"Erro ao extrair informações do PDF: {str(e)}")
        return "", "", 0.0

# Restante do seu código permanece o mesmo...



# Função para iniciar a extração quando o botão for clicado
def iniciar_extracao():
    global pasta_pdf
    pasta_pdf = pasta_entry.get()
    pasta_label.config(text=f"Pasta com PDFs: {pasta_pdf}")

    informacoes_pdf.clear()
    for arquivo in os.listdir(pasta_pdf):
        if arquivo.endswith('.pdf') and not arquivo.startswith('~$'):
            pdf_path = os.path.join(pasta_pdf, arquivo)
            cnpj, razao_social, valor_total = extrair_valores_cnpj_razao_social_e_valor_total(pdf_path)
            informacoes_pdf.append([arquivo, cnpj, razao_social, valor_total])

    # Imprimir dados para diagnóstico
    print("Informações PDF:")
    print(informacoes_pdf)
    print("Colunas:")
    print(colunas)

    # Criar um DataFrame
    df = pd.DataFrame(informacoes_pdf, columns=colunas)

    # Solicitar ao usuário o nome do arquivo Excel
    nome_arquivo_excel = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                      filetypes=[("Arquivos Excel", "*.xlsx")])

    if nome_arquivo_excel:
        # Salvar o DataFrame no arquivo Excel
        df.to_excel(nome_arquivo_excel, index=False, engine='openpyxl')

        # Carregar o arquivo Excel com openpyxl
        wb = load_workbook(nome_arquivo_excel)
        ws = wb.active

        # Adicionar hiperlinks aos nomes dos arquivos (alternativa)
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            cell = row[0]
            nome_arquivo = cell.value
            if nome_arquivo:
                caminho_completo = os.path.join(pasta_pdf, nome_arquivo)
                cell.hyperlink = caminho_completo
                cell.style = "Hyperlink"

        # Salvar o arquivo Excel atualizado
        wb.save(nome_arquivo_excel)

        resultado_label.config(text=f'Valores extraídos e salvos em "{nome_arquivo_excel}".')

# Criar a interface gráfica
root = tk.Tk()
root.title("Extração de CNPJ, Razão Social e Valor Total de PDFs")

# Widgets
instrucoes_label = tk.Label(root, text="Informe o diretório:")
pasta_label = tk.Label(root, text="Pasta com PDFs não selecionada.")
pasta_entry = tk.Entry(root, width=50)
selecionar_button = tk.Button(root, text="Selecionar Pasta",
                              command=lambda: pasta_entry.insert(0, filedialog.askdirectory()))
iniciar_button = tk.Button(root, text="Iniciar Extração", command=iniciar_extracao)
resultado_label = tk.Label(root, text="")

# Layout dos widgets
instrucoes_label.pack()
pasta_entry.pack()
selecionar_button.pack()
pasta_label.pack()
iniciar_button.pack()
resultado_label.pack()

# Lista para armazenar informações dos PDFs
informacoes_pdf = []

# Colunas para o DataFrame
colunas = ["Nome do Arquivo", "CNPJ", "Razão Social", "Valor Total"]

# Executar a aplicação
root.mainloop()
