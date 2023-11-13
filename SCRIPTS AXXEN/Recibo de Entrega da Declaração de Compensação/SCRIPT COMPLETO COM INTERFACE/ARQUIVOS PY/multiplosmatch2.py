import os
import fitz
import re
import pandas as pd
import tkinter as tk
from tkinter import filedialog, Entry
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image


def le_pdf(pdf_path, colunas):
        doc = fitz.open(pdf_path)
        rows = []
        for pagina in doc:
            pagina = pagina.get_text()
            match = re.search(r'([0123]\d\/[01]\d\/\d{4})\n(?=[0123]\d\/[01]\d\/\d{4}\n\d{17}\nCNPJ)', pagina)
            if match:
                periodo = match.group(1)
                valores = re.findall(r'\n(1170|1176|1181|1184|1200|1213|)\n([^\d^\n]*)\n([\d.]*,\d{2})', pagina)
                valores = [list(linha) for linha in valores if all(linha)]
                [linha.insert(0, periodo) for linha in valores]
                rows.extend(valores)
        df = pd.DataFrame(rows, columns=colunas)
        df['Principal'] = df['Principal'].apply(lambda valor: float(valor.replace('.', '').replace(',', '.')))
        return df


def selecionar_pasta():
    global pasta_pdf
    pasta_pdf = filedialog.askdirectory()
    pasta_label.config(text=f"Pasta com PDFs: {pasta_pdf}")


def selecionar_pasta_destino():
    global pasta_destino
    pasta_destino = filedialog.askdirectory()
    pasta_destino_label.config(text=f"Pasta de Destino: {pasta_destino}")


def extrair_e_salvar():
    informacoes_pdf = []
    colunas = ['Período', 'Código', 'Descrição', 'Principal']
    coluna_arquivo = 'Nome do Arquivo'
    df = pd.DataFrame(columns=colunas + [coluna_arquivo])
    for arquivo in os.listdir(pasta_pdf):
        if arquivo.endswith('.pdf'):
            pdf_path = os.path.join(pasta_pdf, arquivo)
            valores_encontrados = le_pdf(pdf_path, colunas)
            row_count, column_count = valores_encontrados.shape
            if row_count > 1 and column_count == 4:
                valores_encontrados[coluna_arquivo] = arquivo
                colunas.append(coluna_arquivo)
            df = pd.concat([df, valores_encontrados], axis=0)

    total_geral = df['Principal'].sum(axis=0)

    qnty_cols = len(df.columns)
    col_idx = list(df.columns).index('Principal')
    ultima_linha = [''] * qnty_cols
    ultima_linha[col_idx] = total_geral
    df.loc['Total Geral'] = ultima_linha

    if not pasta_destino:
        resultado_label.config(text="Por favor, selecione uma pasta de destino.")
        return

    nome_arquivo = nome_arquivo_entry.get().strip()

    if not nome_arquivo:
        resultado_label.config(text="Por favor, insira um nome de arquivo.")
        return

    if not nome_arquivo.endswith(".xlsx"):
        nome_arquivo += ".xlsx"

    excel_filename = os.path.join(pasta_destino, nome_arquivo)

    pasta_completa = os.path.abspath(pasta_pdf)
    df['Nome do Arquivo'] = df.apply(lambda row: f'{os.path.join(pasta_completa, row["Nome do Arquivo"])}', axis=1)

    workbook = Workbook()
    ws = workbook.active

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.hyperlink = cell.value

    workbook.save(excel_filename)
    resultado_label.config(text=f'Valores extraídos e salvos em "{excel_filename}"')


root = tk.Tk()
root.title("Extração de Valores de PDF")

selecionar_button = tk.Button(root, text="Selecionar Pasta com PDFs", command=selecionar_pasta)
selecionar_destino_button = tk.Button(root, text="Selecionar Pasta de Destino", command=selecionar_pasta_destino)
extrair_button = tk.Button(root, text="Extrair e Salvar", command=extrair_e_salvar, width=50)
pasta_label = tk.Label(root, text="Pasta com PDFs não selecionada.")
pasta_destino_label = tk.Label(root, text="Pasta de Destino não selecionada.")
resultado_label = tk.Label(root, text="")

nome_arquivo_label = tk.Label(root, text="Nome do arquivo de saída:")
nome_arquivo_entry = Entry(root)

selecionar_button.pack()
selecionar_destino_button.pack()
pasta_label.pack()
pasta_destino_label.pack()
nome_arquivo_label.pack()
nome_arquivo_entry.pack()
extrair_button.pack()
resultado_label.pack()

root.mainloop()

