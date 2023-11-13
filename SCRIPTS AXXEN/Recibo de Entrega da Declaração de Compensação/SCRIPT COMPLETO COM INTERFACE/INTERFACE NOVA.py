import os
import fitz
import re
import pandas as pd
from typing import Tuple, Any
import tkinter as tk
from tkinter import *
from tkinter import filedialog, Entry
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from glob import glob

from Script_Pronto import pasta_destino_label, nome_arquivo_entry

pasta_pdf = ""
pasta_destino = ""


def extrair_valores_cp(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        texto = ""
        for pagina in doc:
            texto += pagina.get_text()

        padrao = re.compile(
            r'CP TERCEIROS\s+([\d\.]+,\d+)'
            r'|CP SEGURADOS\s+([\d\.]+,\d+)'
            r'|CP PATRONAL\s+([\d\.]+,\d+)'
            r'|IRRF\s+([\d\.]+,\d+)'
            r'|IRPJ\s+([\d\.]+,\d+)'
            r'|CSLL\s+([\d\.]+,\d+)'
            r'|PIS/PASEP\s+([\d\.]+,\d+)'
            r'|COFINS\s+([\d\.]+,\d+)'
            r'|Valor do Pedido:\s+([\d\.]+,\d+)'
            r'|CNPJ:\s*(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})'
            r'|Número do Documento:\s*([\d.-]+)'
            r'|Número da Declaração:\s*([\d.-]+)',
            re.IGNORECASE
        )

        valores_encontrados = padrao.findall(texto)

        cp_terceiros = ""
        cp_segurados = ""
        cp_patronal = ""
        IRRF = ""
        IRPJ = ""
        CSLL = ""
        PASEP = ""
        COFINS = ""
        Valor_do_Pedido = ""
        CNPJ = ""
        Número_do_Documento = ""
        Número_da_Declaração = ""

        for match in valores_encontrados:
            if match[0]:
                cp_terceiros = float(match[0].replace(".", "").replace(",", "."))
            elif match[1]:
                cp_segurados = float(match[1].replace(".", "").replace(",", "."))
            elif match[2]:
                cp_patronal = float(match[2].replace(".", "").replace(",", "."))
            elif match[3]:
                IRRF = float(match[3].replace(".", "").replace(",", "."))
            elif match[4]:
                IRPJ = float(match[4].replace(".", "").replace(",", "."))
            elif match[5]:
                CSLL = float(match[5].replace(".", "").replace(",", "."))
            elif match[6]:
                PASEP = float(match[6].replace(".", "").replace(",", "."))
            elif match[7]:
                COFINS = float(match[7].replace(".", "").replace(",", "."))
            elif match[8]:
                Valor_do_Pedido = float(match[8].replace(".", "").replace(",", "."))
            elif match[9]:
                CNPJ = match[9]
            elif match[10]:
                Número_do_Documento = match[10]
            elif match[11]:
                Número_da_Declaração = match[11]

        return cp_terceiros, cp_segurados, cp_patronal, IRRF, IRPJ, CSLL, PASEP, COFINS, Valor_do_Pedido, CNPJ, Número_do_Documento, Número_da_Declaração
    except Exception as e:
        print(f"Erro ao extrair informações do PDF: {str(e)}")
        return 0.0, 0.0, 0.0  # Retorna 0.0 para os valores numéricos em caso de erro


def selecionar_pasta():
    global pasta_pdf
    pasta_pdf = filedialog.askdirectory()
    pasta_label.config(text=f"Pasta com PDFs: {pasta_pdf}")


def selecionar_pasta_destino():
    global pasta_destino
    pasta_destino = filedialog.askdirectory()
    pasta_destino_label.config(text=f"Pasta de Destino: {pasta_destino}")


def extrair_e_salvar():
    informacoes_pdf = []

    for arquivo in os.listdir(pasta_pdf):
        if arquivo.endswith('.pdf'):
            pdf_path = os.path.join(pasta_pdf, arquivo)
            cp_terceiros, cp_segurados, cp_patronal, IRRF, IRPJ, CSLL, PASEP, COFINS, Valor_do_Pedido, CNPJ, Número_do_Documento, Número_da_Declaração = extrair_valores_cp(
                pdf_path)
            informacoes_pdf.append((arquivo, cp_terceiros, cp_segurados, cp_patronal, IRRF, IRPJ, CSLL, PASEP, COFINS,
                                    Valor_do_Pedido, CNPJ, Número_do_Documento, Número_da_Declaração))

    df = pd.DataFrame(
        informacoes_pdf,
        columns=["Nome do Arquivo", "CP TERCEIROS", "CP SEGURADOS", "CP PATRONAL", "IRRF", "IRPJ", "CSLL", "PIS/PASEP",
                 "COFINS", "Valor do Pedido", "CNPJ", "Número do Documento", "Número da Declaração"]
    )

    colunas_numericas = df.columns[1:]

    df[colunas_numericas] = df[colunas_numericas].apply(pd.to_numeric, errors='coerce').fillna(0)


    df['Soma Total'] = df[colunas_numericas].sum(axis=1)

    total_geral = sum(list(df['Soma Total']))

    qnty_cols = len(df.columns)
    ultima_linha = [''] * qnty_cols
    ultima_linha[-1] = total_geral
    df.loc['Total Geral'] = ultima_linha
    df[colunas_numericas] = df[colunas_numericas].apply(lambda valor: valor.replace(0, ''))



    if not pasta_destino:
        resultado_label.config(text="Por favor, selecione uma pasta de destino.")
        return

    nome_arquivo = nome_arquivo_entry.get().strip()

    if not nome_arquivo:
        resultado_label.config(text="Por favor, insira um nome de arquivo.")
        return

    if not nome_arquivo.endswith(".xlsx"):
        nome_arquivo += ".xlsx"

    excel_filename = os.path.join(pasta_destino, nome_arquivo)

    pasta_completa = os.path.abspath(pasta_pdf)
    df['Nome do Arquivo'] = df.apply(lambda row: f'{os.path.join(pasta_completa, row["Nome do Arquivo"])}', axis=1)

    workbook = Workbook()
    ws = workbook.active

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.hyperlink = cell.value

    workbook.save(excel_filename)
    resultado_label.config(text=f'Valores extraídos e salvos em "{excel_filename}"')



root = tk.Tk()
#root.title("Extração de CNPJ, Razão Social e Valor Total de PDFs")
root.configure(bg="#333333")

instrucoes_label = tk.Label(root, text="Informe o diretório:", font=("Roboto", 15), bg="#333333", fg="white")
pasta_label = tk.Label(root, text="Pasta com PDFs não selecionada.", font=("Roboto", 10), bg="#333333", fg="white")
pasta_entry = tk.Entry(root, width=50, bg="gray", fg="white")
selecionar_button = tk.Button(root, text="Selecionar Pasta", command=lambda: pasta_entry.insert(0, filedialog.askdirectory()), font=("Roboto", 13) ,bg="gray", fg="white")
iniciar_button = tk.Button(root, text="Iniciar Extração", command=selecionar_pasta, font=("Roboto", 13), bg="gray", fg="white")
resultado_label = tk.Label(root, text="")
direitos_autorais = tk.Label(root, text="© 2023 Axxen. Todos os direitos reservados.", font=("Arial", 10), bg="#333333", fg="white")

img = PhotoImage(file = "axxenoficial.png")
label_imagem = Label(root, image=img, bg="#333333").pack()

root.geometry("480x480")
tk.Label(root, text=" ", bg="#333333").pack()
instrucoes_label.pack()
pasta_entry.pack()
tk.Label(root, text="", bg="#333333").pack()
selecionar_button.pack()
tk.Label(root, text="", bg="#333333").pack()
pasta_label.pack()
tk.Label(root, text="", bg="#333333").pack()
iniciar_button.pack()
direitos_autorais.pack(side="bottom")
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
tk.Label(root, text="", bg="#333333").pack()
direitos_autorais.pack(side="bottom")

resultado_label.pack()

root.mainloop()
